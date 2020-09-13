from query import get_database
from constants import attributes
import openpyxl

data = get_database()

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name("Sheet")

# Escribir encabezado
for j, att in enumerate(attributes):
    sheet.cell(row=1, column=j+1).value = att[0]

# Escribir datos y listar por consola tambien
for i, item in enumerate(data):
    for j, att in enumerate(attributes):
        if att[1] in item:
            print(att[0]+": "+str(item[att[1]]), end="\t")
            sheet.cell(row=i+2, column=j+1).value = item[att[1]]
    print(" ")

wb.save("ResumenDatosEscaner.xlsx")